"""
Dynamic Experiment Data Analyzer v3.1
======================================

动态测量数据分析工具

数据源：
1. CSV
2. Excel
3. C++ Sensor Data Backend API

主要功能：
- CSV / Excel / C++ API 数据导入
- Excel Sheet 选择
- 数据列选择
- C++ API timestamp + value 解析
- 数据预览
- 基础统计分析
- 移动平均滤波
- 动态响应分析
- 动态响应曲线

作者：Measurement & Control Project
"""


# ============================================================
# 0. 导入模块
# ============================================================

import csv
import json
import os
import statistics

import tkinter as tk

from tkinter import (
    filedialog,
    messagebox,
    ttk
)

from dataclasses import dataclass

from urllib.request import (
    Request,
    urlopen
)

from urllib.error import (
    URLError,
    HTTPError
)

from datetime import datetime

import matplotlib.pyplot as plt

from openpyxl import load_workbook


# ============================================================
# 1. 实验配置
# ============================================================

@dataclass
class ExperimentConfig:

    # 移动平均窗口
    windows: tuple = (
        3,
        5,
        10,
        20
    )

    # 真实变化点
    true_change_index: int = 101

    # 变化前真实值
    true_value_before: float = 125.0

    # 变化后真实值
    true_value_after: float = 130.0

    # C++ API
    api_url: str = (
        "http://localhost:8080/"
        "sensors/Measurement/data"
    )


CONFIG = ExperimentConfig()


# ============================================================
# 2. GUI 字体
# ============================================================

FONT_NORMAL = (
    "Microsoft YaHei",
    10
)

FONT_TITLE = (
    "Microsoft YaHei",
    15,
    "bold"
)

FONT_HEADER = (
    "Microsoft YaHei",
    11,
    "bold"
)


# ============================================================
# 3. GUI 工具
# ============================================================

class GUIHelper:

    def __init__(
        self,
        root
    ):

        self.root = root

    # --------------------------------------------------------
    # 窗口居中
    # --------------------------------------------------------

    @staticmethod
    def center_window(
        window,
        width,
        height
    ):

        screen_width = (
            window.winfo_screenwidth()
        )

        screen_height = (
            window.winfo_screenheight()
        )

        x = (
            screen_width - width
        ) // 2

        y = (
            screen_height - height
        ) // 2

        window.geometry(
            f"{width}x{height}+{x}+{y}"
        )

    # --------------------------------------------------------
    # 创建窗口
    # --------------------------------------------------------

    def create_window(
        self,
        title,
        width,
        height
    ):

        window = tk.Toplevel(
            self.root
        )

        window.title(
            title
        )

        self.center_window(
            window,
            width,
            height
        )

        window.resizable(
            False,
            False
        )

        return window


# ============================================================
# 4. 数据读取模块
# ============================================================

class DataLoader:

    # --------------------------------------------------------
    # 获取文件扩展名
    # --------------------------------------------------------

    @staticmethod
    def get_extension(
        file_path
    ):

        return os.path.splitext(
            file_path
        )[1].lower()

    # ========================================================
    # CSV
    # ========================================================

    @staticmethod
    def get_csv_headers(
        file_path
    ):

        with open(
            file_path,
            "r",
            encoding="utf-8-sig",
            newline=""
        ) as file:

            reader = csv.DictReader(
                file
            )

            if not reader.fieldnames:

                raise ValueError(
                    "CSV 文件没有找到表头。"
                )

            return list(
                reader.fieldnames
            )

    # --------------------------------------------------------
    # 读取 CSV
    # --------------------------------------------------------

    @staticmethod
    def load_csv(
        file_path,
        column_name
    ):

        data = []

        with open(
            file_path,
            "r",
            encoding="utf-8-sig",
            newline=""
        ) as file:

            reader = csv.DictReader(
                file
            )

            if not reader.fieldnames:

                raise ValueError(
                    "CSV 文件没有找到表头。"
                )

            if column_name not in (
                reader.fieldnames
            ):

                raise ValueError(
                    f"CSV 中没有找到列："
                    f"{column_name}"
                )

            for row in reader:

                value = row.get(
                    column_name
                )

                try:

                    if value not in (
                        None,
                        ""
                    ):

                        data.append(
                            float(value)
                        )

                except (
                    ValueError,
                    TypeError
                ):

                    continue

        return data

    # ========================================================
    # Excel
    # ========================================================

    @staticmethod
    def get_excel_sheets(
        file_path
    ):

        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True
        )

        sheets = (
            workbook.sheetnames
        )

        workbook.close()

        return sheets

    # --------------------------------------------------------
    # Excel 表头
    # --------------------------------------------------------

    @staticmethod
    def get_excel_headers(
        file_path,
        sheet_name
    ):

        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True
        )

        sheet = workbook[
            sheet_name
        ]

        headers = next(
            sheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            ),
            ()
        )

        workbook.close()

        return [
            value
            for value in headers
            if value is not None
        ]

    # --------------------------------------------------------
    # Excel 数据
    # --------------------------------------------------------

    @staticmethod
    def load_excel(
        file_path,
        sheet_name,
        column_name
    ):

        data = []

        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True
        )

        sheet = workbook[
            sheet_name
        ]

        headers = list(
            next(
                sheet.iter_rows(
                    min_row=1,
                    max_row=1,
                    values_only=True
                ),
                ()
            )
        )

        if column_name not in headers:

            workbook.close()

            raise ValueError(
                f"Excel 中没有找到列："
                f"{column_name}"
            )

        column_index = (
            headers.index(
                column_name
            )
        )

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            if (
                column_index
                >= len(row)
            ):

                continue

            value = row[
                column_index
            ]

            try:

                if value is not None:

                    data.append(
                        float(value)
                    )

            except (
                ValueError,
                TypeError
            ):

                continue

        workbook.close()

        return data

    # ========================================================
    # C++ API
    # ========================================================

    @staticmethod
    def load_api(
        url
    ):

        request = Request(

            url,

            headers={
                "Accept":
                "application/json"
            },

            method="GET"

        )

        # ----------------------------------------------------
        # 请求 API
        # ----------------------------------------------------

        try:

            with urlopen(
                request,
                timeout=5
            ) as response:

                raw_data = (
                    response
                    .read()
                    .decode("utf-8")
                )

                result = json.loads(
                    raw_data
                )

        except HTTPError as error:

            raise RuntimeError(
                f"API HTTP 错误："
                f"{error.code}"
            )

        except URLError as error:

            raise RuntimeError(

                "无法连接 C++ 后端。\n\n"

                "请确认：\n"

                "1. SensorDataBackend.exe "
                "正在运行\n"

                "2. API 端口为 8080\n"

                "3. API 地址正确\n\n"

                f"当前地址：{url}"

            ) from error

        except Exception as error:

            raise RuntimeError(
                f"API 数据读取失败："
                f"{error}"
            ) from error

        # ----------------------------------------------------
        # 解析数据
        # ----------------------------------------------------

        records = []

        api_data = result.get(
            "data",
            []
        )

        for item in api_data:

            try:

                # --------------------------------------------
                # 标准格式
                #
                # {
                #     "timestamp": "...",
                #     "value": 125.36
                # }
                # --------------------------------------------

                if isinstance(
                    item,
                    dict
                ):

                    value = float(
                        item["value"]
                    )

                    timestamp = (
                        item.get(
                            "timestamp"
                        )
                    )

                # --------------------------------------------
                # 兼容：
                #
                # {
                #     "data": [125.3, 126.4]
                # }
                # --------------------------------------------

                else:

                    value = float(
                        item
                    )

                    timestamp = None

                records.append({

                    "timestamp":
                        timestamp,

                    "value":
                        value

                })

            except (
                KeyError,
                ValueError,
                TypeError
            ):

                continue

        # ----------------------------------------------------
        # 检查
        # ----------------------------------------------------

        if not records:

            raise ValueError(
                "API 没有返回有效测量数据。"
            )

        return records


# ============================================================
# 5. 通用列表选择窗口
# ============================================================

def select_from_list(
    gui,
    title,
    description,
    items,
    width=600,
    height=450
):

    result = {
        "value": None
    }

    window = gui.create_window(
        title,
        width,
        height
    )

    tk.Label(
        window,
        text=title,
        font=FONT_TITLE
    ).pack(
        pady=(20, 5)
    )

    tk.Label(
        window,
        text=description,
        font=FONT_NORMAL
    ).pack(
        pady=(0, 15)
    )

    frame = tk.Frame(
        window
    )

    frame.pack(
        fill="both",
        expand=True,
        padx=25
    )

    scrollbar = ttk.Scrollbar(
        frame,
        orient="vertical"
    )

    scrollbar.pack(
        side="right",
        fill="y"
    )

    tree = ttk.Treeview(

        frame,

        columns=(
            "value",
        ),

        show="headings",

        selectmode="browse",

        yscrollcommand=(
            scrollbar.set
        )

    )

    tree.heading(
        "value",
        text="名称"
    )

    tree.column(
        "value",
        width=480
    )

    tree.pack(
        side="left",
        fill="both",
        expand=True
    )

    scrollbar.config(
        command=tree.yview
    )

    for item in items:

        tree.insert(
            "",
            tk.END,
            values=(
                str(item),
            )
        )

    if items:

        first = (
            tree.get_children()[0]
        )

        tree.selection_set(
            first
        )

        tree.focus(
            first
        )

    # --------------------------------------------------------
    # 确定
    # --------------------------------------------------------

    def confirm():

        selection = (
            tree.selection()
        )

        if not selection:

            messagebox.showwarning(

                "提示",

                "请选择一个项目。",

                parent=window

            )

            return

        result["value"] = (
            tree.item(
                selection[0],
                "values"
            )[0]
        )

        window.destroy()

    # --------------------------------------------------------
    # 取消
    # --------------------------------------------------------

    def cancel():

        result["value"] = None

        window.destroy()

    tree.bind(
        "<Double-Button-1>",
        lambda event: confirm()
    )

    window.bind(
        "<Return>",
        lambda event: confirm()
    )

    window.bind(
        "<Escape>",
        lambda event: cancel()
    )

    buttons = tk.Frame(
        window
    )

    buttons.pack(
        pady=20
    )

    tk.Button(
        buttons,
        text="确定",
        command=confirm,
        width=12,
        font=FONT_NORMAL
    ).pack(
        side="left",
        padx=10
    )

    tk.Button(
        buttons,
        text="取消",
        command=cancel,
        width=12,
        font=FONT_NORMAL
    ).pack(
        side="left",
        padx=10
    )

    window.grab_set()

    window.focus_force()

    gui.root.wait_window(
        window
    )

    return result["value"]


# ============================================================
# 6. 数据源选择
# ============================================================

def select_data_source(
    gui
):

    result = {
        "value": None
    }

    window = gui.create_window(
        "选择数据源",
        560,
        430
    )

    tk.Label(
        window,
        text="选择数据源",
        font=FONT_TITLE
    ).pack(
        pady=(30, 10)
    )

    tk.Label(
        window,
        text="请选择本次实验的数据来源：",
        font=FONT_NORMAL
    ).pack(
        pady=(0, 25)
    )

    source = tk.StringVar(
        value="Excel"
    )

    options = [

        (
            "CSV 文件",
            "CSV"
        ),

        (
            "Excel 文件",
            "Excel"
        ),

        (
            "C++ 实时 API",
            "API"
        )

    ]

    for text, value in options:

        tk.Radiobutton(

            window,

            text=text,

            variable=source,

            value=value,

            font=FONT_NORMAL

        ).pack(

            anchor="w",

            padx=130,

            pady=8

        )

    # --------------------------------------------------------
    # 确定
    # --------------------------------------------------------

    def confirm():

        result["value"] = (
            source.get()
        )

        window.destroy()

    # --------------------------------------------------------
    # 取消
    # --------------------------------------------------------

    def cancel():

        result["value"] = None

        window.destroy()

    buttons = tk.Frame(
        window
    )

    buttons.pack(
        pady=30
    )

    tk.Button(
        buttons,
        text="确定",
        command=confirm,
        width=14,
        font=FONT_NORMAL
    ).pack(
        side="left",
        padx=10
    )

    tk.Button(
        buttons,
        text="取消",
        command=cancel,
        width=14,
        font=FONT_NORMAL
    ).pack(
        side="left",
        padx=10
    )

    window.bind(
        "<Return>",
        lambda e: confirm()
    )

    window.bind(
        "<Escape>",
        lambda e: cancel()
    )

    window.grab_set()

    window.focus_force()

    gui.root.wait_window(
        window
    )

    return result["value"]


# ============================================================
# 7. 实验参数
# ============================================================

def experiment_parameter_window(
    gui,
    config
):

    result = {
        "confirmed": False
    }

    window = gui.create_window(
        "实验参数设置",
        600,
        520
    )

    tk.Label(
        window,
        text="实验参数设置",
        font=FONT_TITLE
    ).pack(
        pady=(25, 5)
    )

    tk.Label(
        window,
        text="请输入本次实验的真实参数",
        font=FONT_NORMAL
    ).pack(
        pady=(0, 20)
    )

    frame = tk.Frame(
        window
    )

    frame.pack(
        padx=60,
        fill="x"
    )

    before_var = tk.StringVar(
        value=str(
            config.true_value_before
        )
    )

    after_var = tk.StringVar(
        value=str(
            config.true_value_after
        )
    )

    change_var = tk.StringVar(
        value=str(
            config.true_change_index
        )
    )

    windows_var = tk.StringVar(

        value=",".join(

            map(
                str,
                config.windows
            )

        )

    )

    fields = [

        (
            "变化前真实值：",
            before_var,
            "mm"
        ),

        (
            "变化后真实值：",
            after_var,
            "mm"
        ),

        (
            "真实变化点：",
            change_var,
            "sample"
        ),

        (
            "滤波窗口：",
            windows_var,
            "例如 3,5,10,20"
        )

    ]

    for row, (
        label,
        variable,
        unit
    ) in enumerate(fields):

        tk.Label(
            frame,
            text=label,
            font=FONT_HEADER
        ).grid(
            row=row,
            column=0,
            sticky="w",
            pady=10
        )

        tk.Entry(
            frame,
            textvariable=variable,
            width=20,
            font=FONT_NORMAL
        ).grid(
            row=row,
            column=1,
            padx=10
        )

        tk.Label(
            frame,
            text=unit,
            font=FONT_NORMAL
        ).grid(
            row=row,
            column=2
        )

    # --------------------------------------------------------
    # 确定
    # --------------------------------------------------------

    def confirm():

        try:

            before = float(
                before_var.get()
            )

            after = float(
                after_var.get()
            )

            change_index = int(
                change_var.get()
            )

            windows = tuple(

                int(
                    x.strip()
                )

                for x in
                windows_var.get().split(",")

                if x.strip()

            )

            if not windows:

                raise ValueError(
                    "至少需要一个滤波窗口。"
                )

            if any(
                x <= 0
                for x in windows
            ):

                raise ValueError(
                    "滤波窗口必须大于 0。"
                )

            if change_index < 1:

                raise ValueError(
                    "变化点必须大于 0。"
                )

            config.true_value_before = (
                before
            )

            config.true_value_after = (
                after
            )

            config.true_change_index = (
                change_index
            )

            config.windows = (
                windows
            )

            result["confirmed"] = True

            window.destroy()

        except ValueError as error:

            messagebox.showerror(
                "参数错误",
                str(error),
                parent=window
            )

    # --------------------------------------------------------
    # 取消
    # --------------------------------------------------------

    def cancel():

        window.destroy()

    buttons = tk.Frame(
        window
    )

    buttons.pack(
        pady=30
    )

    tk.Button(
        buttons,
        text="确定",
        command=confirm,
        width=14,
        font=FONT_NORMAL
    ).pack(
        side="left",
        padx=10
    )

    tk.Button(
        buttons,
        text="取消",
        command=cancel,
        width=14,
        font=FONT_NORMAL
    ).pack(
        side="left",
        padx=10
    )

    window.bind(
        "<Return>",
        lambda e: confirm()
    )

    window.bind(
        "<Escape>",
        lambda e: cancel()
    )

    window.grab_set()

    window.focus_force()

    gui.root.wait_window(
        window
    )

    return result["confirmed"]


# ============================================================
# 8. 数据预览
# ============================================================

def preview_data(
    gui,
    source_name,
    file_path,
    sheet_name,
    column_name,
    data,
    timestamps=None
):

    result = {
        "continue": False
    }

    if timestamps is None:

        timestamps = []

    window = gui.create_window(
        "数据预览",
        850,
        680
    )

    tk.Label(
        window,
        text="测量数据预览",
        font=FONT_TITLE
    ).pack(
        pady=(20, 10)
    )

    # --------------------------------------------------------
    # 文件信息
    # --------------------------------------------------------

    if source_name == "API":

        info = (

            "数据源：C++ Sensor Data Backend\n"

            f"API：{CONFIG.api_url}\n"

            f"样本数量：{len(data)}"

        )

    else:

        info = (

            f"文件："
            f"{os.path.basename(file_path)}\n"

            f"Sheet：{sheet_name}\n"

            f"数据列：{column_name}\n"

            f"数据数量：{len(data)}"

        )

    tk.Label(
        window,
        text=info,
        font=FONT_NORMAL,
        justify="left",
        anchor="w"
    ).pack(
        anchor="w",
        padx=30,
        pady=(0, 15)
    )

    # --------------------------------------------------------
    # 表格
    # --------------------------------------------------------

    frame = tk.Frame(
        window
    )

    frame.pack(
        fill="both",
        expand=True,
        padx=30
    )

    scrollbar = ttk.Scrollbar(
        frame,
        orient="vertical"
    )

    scrollbar.pack(
        side="right",
        fill="y"
    )

    if source_name == "API":

        columns = (
            "index",
            "timestamp",
            "value"
        )

        tree = ttk.Treeview(

            frame,

            columns=columns,

            show="headings",

            yscrollcommand=(
                scrollbar.set
            )

        )

        tree.heading(
            "index",
            text="序号"
        )

        tree.heading(
            "timestamp",
            text="时间戳"
        )

        tree.heading(
            "value",
            text="测量值"
        )

        tree.column(
            "index",
            width=70,
            anchor="center"
        )

        tree.column(
            "timestamp",
            width=220,
            anchor="center"
        )

        tree.column(
            "value",
            width=180,
            anchor="center"
        )

        for index, value in enumerate(
            data[:100],
            start=1
        ):

            timestamp = ""

            if (
                index - 1
                < len(timestamps)
            ):

                timestamp = (
                    timestamps[
                        index - 1
                    ]
                    or ""
                )

            tree.insert(
                "",
                tk.END,
                values=(
                    index,
                    timestamp,
                    f"{value:.6f}"
                )
            )

    else:

        columns = (
            "index",
            "value"
        )

        tree = ttk.Treeview(

            frame,

            columns=columns,

            show="headings",

            yscrollcommand=(
                scrollbar.set
            )

        )

        tree.heading(
            "index",
            text="序号"
        )

        tree.heading(
            "value",
            text=column_name
        )

        tree.column(
            "index",
            width=100,
            anchor="center"
        )

        tree.column(
            "value",
            width=520,
            anchor="center"
        )

        for index, value in enumerate(
            data[:100],
            start=1
        ):

            tree.insert(
                "",
                tk.END,
                values=(
                    index,
                    f"{value:.6f}"
                )
            )

    tree.pack(
        side="left",
        fill="both",
        expand=True
    )

    scrollbar.config(
        command=tree.yview
    )

    # --------------------------------------------------------
    # 统计
    # --------------------------------------------------------

    mean_value = (
        statistics.mean(data)
    )

    tk.Label(

        window,

        text=(

            f"最小值："
            f"{min(data):.4f}    "

            f"最大值："
            f"{max(data):.4f}    "

            f"平均值："
            f"{mean_value:.4f}"

        ),

        font=FONT_HEADER

    ).pack(
        pady=15
    )

    # --------------------------------------------------------
    # 操作
    # --------------------------------------------------------

    def start():

        result["continue"] = True

        window.destroy()

    def cancel():

        result["continue"] = False

        window.destroy()

    buttons = tk.Frame(
        window
    )

    buttons.pack(
        pady=10
    )

    tk.Button(
        buttons,
        text="开始分析",
        command=start,
        width=12,
        font=FONT_NORMAL
    ).pack(
        side="left",
        padx=10
    )

    tk.Button(
        buttons,
        text="取消",
        command=cancel,
        width=12,
        font=FONT_NORMAL
    ).pack(
        side="left",
        padx=10
    )

    window.bind(
        "<Return>",
        lambda e: start()
    )

    window.bind(
        "<Escape>",
        lambda e: cancel()
    )

    window.grab_set()

    window.focus_force()

    gui.root.wait_window(
        window
    )

    return result["continue"]


# ============================================================
# 9. 数据分析
# ============================================================

class DataAnalyzer:

    def __init__(
        self,
        config
    ):

        self.config = config

    # --------------------------------------------------------
    # 基础分析
    # --------------------------------------------------------

    def calculate(
        self,
        data
    ):

        if not data:

            raise ValueError(
                "没有可分析的数据。"
            )

        mean_value = (
            statistics.mean(data)
        )

        if len(data) > 1:

            std_value = (
                statistics.stdev(data)
            )

        else:

            std_value = 0.0

        minimum = min(data)

        maximum = max(data)

        peak_to_peak = (
            maximum - minimum
        )

        before = (
            self.config
            .true_value_before
        )

        after = (
            self.config
            .true_value_after
        )

        change = (
            after - before
        )

        change_index = (
            self.config
            .true_change_index
        )

        # ----------------------------------------------------
        # 最大绝对误差
        # ----------------------------------------------------

        maximum_error = 0.0

        for index, value in enumerate(
            data
        ):

            reference = (

                before

                if index < change_index

                else after

            )

            maximum_error = max(

                maximum_error,

                abs(
                    value - reference
                )

            )

        # ----------------------------------------------------
        # 稳态误差
        # ----------------------------------------------------

        steady_count = max(

            1,

            int(
                len(data) * 0.1
            )

        )

        steady_data = data[
            -steady_count:
        ]

        steady_mean = (
            statistics.mean(
                steady_data
            )
        )

        steady_error = abs(

            steady_mean
            -
            after

        )

        # ----------------------------------------------------
        # 90% 响应时间
        # ----------------------------------------------------

        target = (

            before

            +

            change * 0.9

        )

        response_index = None

        start = max(
            0,
            change_index - 1
        )

        for index in range(
            start,
            len(data)
        ):

            if data[index] >= target:

                response_index = (
                    index + 1
                )

                break

        response_time = None

        if response_index is not None:

            response_time = (

                response_index

                -

                change_index

            )

        return {

            "sample_count":
                len(data),

            "mean":
                mean_value,

            "std":
                std_value,

            "minimum":
                minimum,

            "maximum":
                maximum,

            "peak_to_peak":
                peak_to_peak,

            "change_value":
                change,

            "maximum_error":
                maximum_error,

            "steady_mean":
                steady_mean,

            "steady_error":
                steady_error,

            "response_time":
                response_time

        }


# ============================================================
# 10. 移动平均
# ============================================================

def moving_average(
    data,
    window
):

    if window <= 0:

        raise ValueError(
            "窗口必须大于 0。"
        )

    result = []

    buffer = []

    running_sum = 0.0

    for value in data:

        buffer.append(
            value
        )

        running_sum += value

        if len(buffer) > window:

            running_sum -= (
                buffer.pop(0)
            )

        result.append(

            running_sum
            /
            len(buffer)

        )

    return result


# ============================================================
# 11. 分析结果窗口
# ============================================================

def show_results(
    gui,
    stats
):

    window = gui.create_window(
        "数据分析结果",
        650,
        650
    )

    tk.Label(
        window,
        text="测量数据分析结果",
        font=FONT_TITLE
    ).pack(
        pady=(20, 20)
    )

    frame = tk.Frame(
        window
    )

    frame.pack(
        fill="both",
        expand=True,
        padx=40
    )

    results = [

        (
            "样本数量",
            f"{stats['sample_count']}"
        ),

        (
            "平均值",
            f"{stats['mean']:.4f} mm"
        ),

        (
            "标准差",
            f"{stats['std']:.4f} mm"
        ),

        (
            "最小值",
            f"{stats['minimum']:.4f} mm"
        ),

        (
            "最大值",
            f"{stats['maximum']:.4f} mm"
        ),

        (
            "峰峰值",
            f"{stats['peak_to_peak']:.4f} mm"
        ),

        (
            "真实变化量",
            f"{stats['change_value']:.4f} mm"
        ),

        (
            "最大绝对误差",
            f"{stats['maximum_error']:.4f} mm"
        ),

        (
            "稳态平均值",
            f"{stats['steady_mean']:.4f} mm"
        ),

        (
            "稳态误差",
            f"{stats['steady_error']:.4f} mm"
        )

    ]

    if (
        stats["response_time"]
        is None
    ):

        response_text = (
            "未达到 90%"
        )

    else:

        response_text = (

            f"{stats['response_time']} "
            f"samples"

        )

    results.append(

        (
            "90%响应时间",
            response_text
        )

    )

    for name, value in results:

        row = tk.Frame(
            frame
        )

        row.pack(
            fill="x",
            pady=6
        )

        tk.Label(
            row,
            text=name,
            font=FONT_HEADER,
            width=18,
            anchor="w"
        ).pack(
            side="left"
        )

        tk.Label(
            row,
            text=value,
            font=FONT_NORMAL,
            anchor="w"
        ).pack(
            side="left"
        )

    def close():

        window.destroy()

    tk.Button(
        window,
        text="查看动态响应曲线",
        command=close,
        width=20,
        font=FONT_NORMAL
    ).pack(
        pady=20
    )

    window.bind(
        "<Return>",
        lambda e: close()
    )

    window.bind(
        "<Escape>",
        lambda e: close()
    )

    window.grab_set()

    window.focus_force()

    gui.root.wait_window(
        window
    )


# ============================================================
# 12. 动态响应曲线
# ============================================================

def plot_dynamic_response(
    data,
    config,
    timestamps=None
):

    if timestamps is None:

        timestamps = []

    plt.figure(
        figsize=(10, 6)
    )

    # --------------------------------------------------------
    # X轴
    # --------------------------------------------------------

    x = range(
        1,
        len(data) + 1
    )

    # --------------------------------------------------------
    # 原始数据
    # --------------------------------------------------------

    plt.plot(

        x,

        data,

        label="Raw Data",

        alpha=0.45

    )

    # --------------------------------------------------------
    # 移动平均
    # --------------------------------------------------------

    for window in (
        config.windows
    ):

        filtered = moving_average(

            data,

            window

        )

        plt.plot(

            x,

            filtered,

            label=f"Window {window}"

        )

    # --------------------------------------------------------
    # 真实变化点
    # --------------------------------------------------------

    plt.axvline(

        config.true_change_index,

        linestyle="--",

        label="True Change"

    )

    # --------------------------------------------------------
    # 真实值
    # --------------------------------------------------------

    plt.axhline(

        config.true_value_before,

        linestyle=":"

    )

    plt.axhline(

        config.true_value_after,

        linestyle=":"

    )

    # --------------------------------------------------------
    # 标签
    # --------------------------------------------------------

    plt.xlabel(
        "Measurement Index"
    )

    plt.ylabel(
        "Distance (mm)"
    )

    plt.title(
        "Dynamic Response of Moving Average Filters"
    )

    plt.grid(
        True
    )

    plt.legend()

    plt.tight_layout()

    plt.show()


# ============================================================
# 13. 主程序
# ============================================================

def main():

    root = tk.Tk()

    root.withdraw()

    gui = GUIHelper(
        root
    )

    try:

        # ====================================================
        # 选择数据源
        # ====================================================

        source = select_data_source(
            gui
        )

        if not source:

            return

        file_path = ""

        selected_sheet = ""

        selected_column = ""

        timestamps = []

        # ====================================================
        # CSV
        # ====================================================

        if source == "CSV":

            file_path = (
                filedialog
                .askopenfilename(

                    title="选择 CSV 数据文件",

                    filetypes=[
                        (
                            "CSV 文件",
                            "*.csv"
                        )
                    ]

                )
            )

            if not file_path:

                return

            headers = (
                DataLoader
                .get_csv_headers(
                    file_path
                )
            )

            if (
                "Measurement_mm"
                in headers
            ):

                selected_column = (
                    "Measurement_mm"
                )

            else:

                selected_column = (
                    select_from_list(

                        gui,

                        "选择数据列",

                        "请选择用于分析的数据列：",

                        headers

                    )
                )

            if not selected_column:

                return

            selected_sheet = "CSV"

            data = (
                DataLoader
                .load_csv(

                    file_path,

                    selected_column

                )
            )

        # ====================================================
        # Excel
        # ====================================================

        elif source == "Excel":

            file_path = (
                filedialog
                .askopenfilename(

                    title="选择 Excel 数据文件",

                    filetypes=[
                        (
                            "Excel 文件",
                            "*.xlsx"
                        )
                    ]

                )
            )

            if not file_path:

                return

            sheets = (
                DataLoader
                .get_excel_sheets(
                    file_path
                )
            )

            selected_sheet = (
                select_from_list(

                    gui,

                    "选择工作表",

                    "请选择需要分析的 Excel 工作表：",

                    sheets

                )
            )

            if not selected_sheet:

                return

            headers = (
                DataLoader
                .get_excel_headers(

                    file_path,

                    selected_sheet

                )
            )

            if not headers:

                raise ValueError(
                    "当前工作表没有有效表头。"
                )

            selected_column = (
                select_from_list(

                    gui,

                    "选择数据列",

                    "请选择用于分析的数据列：",

                    headers

                )
            )

            if not selected_column:

                return

            data = (
                DataLoader
                .load_excel(

                    file_path,

                    selected_sheet,

                    selected_column

                )
            )

        # ====================================================
        # C++ API
        # ====================================================

        else:

            records = (
                DataLoader
                .load_api(

                    CONFIG.api_url

                )
            )

            # -----------------------------------------------
            # 提取数值
            # -----------------------------------------------

            data = [

                item["value"]

                for item in records

            ]

            # -----------------------------------------------
            # 提取时间戳
            # -----------------------------------------------

            timestamps = [

                item["timestamp"]

                for item in records

            ]

            selected_sheet = (
                "C++ API"
            )

            selected_column = (
                "Measurement"
            )

        # ====================================================
        # 数据检查
        # ====================================================

        if not data:

            raise ValueError(
                "没有读取到有效的数值数据。"
            )

        print()

        print(
            "=" * 65
        )

        print(
            "Dynamic Experiment "
            "Data Analyzer v3.1"
        )

        print(
            "=" * 65
        )

        print(
            f"数据源：{source}"
        )

        if file_path:

            print(
                f"文件："
                f"{os.path.basename(file_path)}"
            )

        print(
            f"Sheet："
            f"{selected_sheet}"
        )

        print(
            f"数据列："
            f"{selected_column}"
        )

        print(
            f"样本数量："
            f"{len(data)}"
        )

        # ----------------------------------------------------
        # API 时间戳
        # ----------------------------------------------------

        if timestamps:

            print(
                f"第一条时间："
                f"{timestamps[0]}"
            )

            print(
                f"最后一条时间："
                f"{timestamps[-1]}"
            )

        print(
            "=" * 65
        )

        # ====================================================
        # 数据预览
        # ====================================================

        if not preview_data(

            gui,

            source,

            file_path,

            selected_sheet,

            selected_column,

            data,

            timestamps

        ):

            return

        # ====================================================
        # 实验参数
        # ====================================================

        if not experiment_parameter_window(

            gui,

            CONFIG

        ):

            return

        # ====================================================
        # 数据分析
        # ====================================================

        analyzer = DataAnalyzer(
            CONFIG
        )

        statistics_result = (
            analyzer.calculate(
                data
            )
        )

        # ====================================================
        # 显示结果
        # ====================================================

        show_results(

            gui,

            statistics_result

        )

        # ====================================================
        # 动态曲线
        # ====================================================

        plot_dynamic_response(

            data,

            CONFIG,

            timestamps

        )

    except Exception as error:

        messagebox.showerror(

            "程序错误",

            str(error)

        )

        print()

        print(
            "ERROR:"
        )

        print(
            error
        )

    finally:

        root.destroy()


# ============================================================
# 14. 程序入口
# ============================================================

if __name__ == "__main__":

    main()